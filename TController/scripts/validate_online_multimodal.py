#!/usr/bin/env python3
"""Replay a titration workbook through the causal TController detector.

By default rows are aggregated to one mean sample per volume, which is convenient
but hides a production path: the firmware reports a spectrum per AS7341 frame while
the volume comes from the pump, so several spectra share one volume and the tracker
must hold its volume-normalised speed instead of normalising by a zero step.
Aggregated replays report ``repeated_spectral_volume=0`` and therefore never
exercise that path.  Pass ``--raw-frames`` to feed every row at its own volume.
"""

from __future__ import annotations

import argparse
import sys
from collections.abc import Iterator
from pathlib import Path
from typing import Any

import numpy as np
import openpyxl

PROJ = Path(__file__).resolve().parents[1]
SRC = PROJ / "src"
sys.path.insert(0, str(SRC))

from DataProcessor.endpoint import EndpointDetector

# One replay event: volume, potential in ADC LSB (or None), spectrum (or None).
Event = tuple[float, float | None, np.ndarray | None]


def load_rows(path: Path) -> tuple[list[tuple[float, float]], list[tuple[float, np.ndarray]]]:
    """Return the raw (volume, potential) and (volume, spectrum) rows, duplicates kept."""
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    potential_rows: list[tuple[float, float]] = []
    spectrum_rows: list[tuple[float, np.ndarray]] = []
    for row in workbook["电位-体积曲线"].iter_rows(min_row=2, values_only=True):
        if row[0] is not None and row[1] is not None:
            potential_rows.append((round(float(str(row[0])), 6), float(str(row[1]))))
    for row in workbook["光谱数据"].iter_rows(min_row=2, values_only=True):
        if row[0] is not None and all(value is not None for value in row[1:9]):
            spectrum_rows.append(
                (
                    round(float(str(row[0])), 6),
                    np.asarray([float(str(value)) for value in row[1:9]], dtype=np.float64),
                )
            )
    workbook.close()
    return potential_rows, spectrum_rows


def _aggregated_events(
    potential_rows: list[tuple[float, float]], spectrum_rows: list[tuple[float, np.ndarray]]
) -> Iterator[Event]:
    """One mean potential and one mean spectrum per shared volume."""
    potentials: dict[float, list[float]] = {}
    spectra: dict[float, list[np.ndarray]] = {}
    for volume, potential in potential_rows:
        potentials.setdefault(volume, []).append(potential)
    for volume, spectrum in spectrum_rows:
        spectra.setdefault(volume, []).append(spectrum)
    for volume in sorted(set(potentials) & set(spectra)):
        yield (
            volume,
            float(np.mean(potentials[volume])),
            np.mean(spectra[volume], axis=0),
        )


def _raw_events(
    potential_rows: list[tuple[float, float]], spectrum_rows: list[tuple[float, np.ndarray]]
) -> Iterator[Event]:
    """Every row at its own volume, merged on volume so both streams stay causal."""
    potentials = sorted(potential_rows, key=lambda row: row[0])
    spectra = sorted(spectrum_rows, key=lambda row: row[0])
    index = 0
    for volume, spectrum in spectra:
        while index < len(potentials) and potentials[index][0] <= volume:
            yield (potentials[index][0], potentials[index][1], None)
            index += 1
        yield (volume, None, spectrum)
    for pot_volume, potential in potentials[index:]:
        yield (pot_volume, potential, None)


def replay(path: Path, flow_rate: float, raw_frames: bool = False) -> dict[str, Any]:
    potential_rows, spectrum_rows = load_rows(path)
    builder = _raw_events if raw_frames else _aggregated_events
    detector = EndpointDetector(flow_rate=flow_rate)
    feature_rows: list[dict[str, Any]] = []
    potential_scale = 3.3 / 65535.0
    volumes: list[float] = []
    for volume, potential, spectrum in builder(potential_rows, spectrum_rows):
        if potential is not None:
            # The detector's production contract is volts; workbooks store ADC LSB.
            detector.feed_potential(
                volume, volume / flow_rate, potential * potential_scale - 1.1
            )
        if spectrum is not None:
            detector.feed_spectrum(volume, spectrum)
            feature_rows.append(detector.diagnostics())
            volumes.append(volume)

    spectral_rows = [row["spectral_features"] for row in feature_rows]
    valid = [row for row in spectral_rows if row.get("valid_frame")]
    result = detector.detect()
    return {
        "path": str(path),
        "samples": len(volumes),
        "volume_range": (min(volumes), max(volumes)) if volumes else (0.0, 0.0),
        "result": result,
        "detector": detector,
        "spectral_rows": valid,
    }


def _peak(rows: list[dict[str, Any]], key: str) -> tuple[float, float] | None:
    values = [(float(row["volume"]), float(row.get(key, 0.0))) for row in rows if row.get("volume") is not None]
    return max(values, key=lambda item: item[1]) if values else None


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", type=Path, required=True, help="Workbook with the two titration sheets")
    parser.add_argument("--flow-rate", type=float, default=0.00603752, help="Pump flow in mL/s")
    parser.add_argument(
        "--raw-frames",
        action="store_true",
        help="Feed every row separately instead of one mean per volume, so repeated "
        "volumes reach the tracker the way they do in production",
    )
    args = parser.parse_args()
    report = replay(args.input, args.flow_rate, raw_frames=args.raw_frames)
    result = report["result"]
    rows = report["spectral_rows"]
    quality = (result or {}).get("reliability", {}).get("data_quality", {})
    print(f"input={report['path']} raw_frames={args.raw_frames}")
    print(f"samples={report['samples']} volume={report['volume_range'][0]:.6f}..{report['volume_range'][1]:.6f} mL")
    print(
        f"repeated_volume={quality.get('repeated_spectral_volume')} "
        f"nonmonotonic={quality.get('nonmonotonic_volume')}"
    )
    print(f"js_speed_peak={_peak(rows, 'js_speed')}")
    print(f"cross_curvature_peak={_peak(rows, 'cross_curvature')}")
    print(f"potential_state={report['detector'].potential_state} spectral_state={report['detector'].spectral_state}")
    print(f"result={result}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
