#!/usr/bin/env python3
"""TController 滴定终点检测算法验证（高效版）。"""
from __future__ import annotations

import sys
import warnings
from pathlib import Path

import numpy as np
import openpyxl

warnings.filterwarnings("ignore")

PROJ = Path(__file__).resolve().parents[1]
SRC = PROJ / "src"
sys.path.insert(0, str(SRC))

from DataProcessor.calibration import FLOW_RATE, update_from_file

DATA_FILE = Path(
    "/home/zhiyir/文档/xwechat_files/wxid_l267qu0nkh512_1601/msg/attach/"
    "d819785e5916791e7d3b8b1199d7af8d/2026-06/Rec/"
    "da42816e8394beff/F/1/titration_result.xlsx"
)
OUT_DIR = PROJ / "data" / "validation"
OUT_DIR.mkdir(parents=True, exist_ok=True)

update_from_file()
print(f"流速: {FLOW_RATE:.6f} mL/s")

# ═══════════════════════════════════════════════════════════════════════
#  1. Load & preprocess
# ═══════════════════════════════════════════════════════════════════════

print("\n[1] 加载数据…")
wb = openpyxl.load_workbook(str(DATA_FILE), read_only=True)
rows_p = list(wb["电位-体积曲线"].iter_rows(min_row=2, values_only=True))
rows_s = list(wb["光谱数据"].iter_rows(min_row=2, values_only=True))
wb.close()

pot_vol = np.array([r[0] for r in rows_p], dtype=np.float64)
pot_val = np.array([r[1] for r in rows_p], dtype=np.float64)
spec_vol = np.array([r[0] for r in rows_s], dtype=np.float64)
spec_8 = np.array([list(r[1:9]) for r in rows_s], dtype=np.float64)

def group_mean(vol, vals):
    u, inv, cnt = np.unique(np.round(vol, 6), return_inverse=True, return_counts=True)
    if vals.ndim == 1:
        m = np.bincount(inv, weights=vals) / cnt
    else:
        m = np.column_stack([np.bincount(inv, weights=vals[:, c]) / cnt for c in range(vals.shape[1])])
    return u, m

uvol_p, mpot = group_mean(pot_vol, pot_val)
uvol_s, mspec = group_mean(spec_vol, spec_8)
print(f"  电位: {len(uvol_p)} 步 ({uvol_p[0]:.4f}–{uvol_p[-1]:.4f} mL)")
print(f"  光谱: {len(uvol_s)} 步")

print("[2] 对齐 & 重建光谱…")
mspec_interp = np.column_stack([np.interp(uvol_p, uvol_s, mspec[:, c]) for c in range(8)])

calib = np.load(str(PROJ / "data" / "calibre.npz"), allow_pickle=True)
mat, ofs, fac = calib["spectral_matrix"], calib["spectral_offsets"], calib["spectral_factors"]
full = np.zeros((len(mspec_interp), 10), dtype=np.float64)
full[:, :8] = mspec_interp
corrected = fac * np.maximum(full - ofs, 0.0)
spectra_721 = corrected @ mat.T
print(f"  全光谱: {spectra_721.shape} ({np.sum(spectra_721<0)}/{spectra_721.size} 负值)")

t = uvol_p / FLOW_RATE
N = len(uvol_p)

# ═══════════════════════════════════════════════════════════════════════
#  2. AMPD analysis (efficient with limited max_scale)
# ═══════════════════════════════════════════════════════════════════════

print("\n[3] 电位通道分析 (savgol w=15, AMPD max_scale=200)…")
from DataProcessor.endpoint import savgol_filter


def ampd_limited(signal, max_scale=200):
    Ns = len(signal)
    if Ns < 4: return np.array([], dtype=int)
    L = min(max_scale, Ns // 2)
    LMS = np.ones((L - 1, Ns), dtype=np.int32)
    for k in range(2, L + 1):
        r = np.ones(Ns, dtype=np.int32)
        for i in range(k, Ns - k):
            if signal[i] > signal[i - k] and signal[i] >= signal[i + k]:
                r[i] = 0
        LMS[k - 2] = r
    row_sums = LMS.sum(axis=1)
    # Find best scale from first 25% to avoid edge-effect artifacts
    search = max(1, int((L - 1) * 0.25))
    best_k = int(np.argmin(row_sums[:search])) + 2
    peaks = np.where(LMS[best_k - 2] == 0)[0]
    return peaks

dv = np.diff(mpot)
dt = np.diff(t)
with np.errstate(divide="ignore", invalid="ignore"):
    deriv = np.where(dt > 0, dv / dt, 0.0)
deriv_sm = savgol_filter(deriv, window=15, order=2)
peaks_ampd = ampd_limited(deriv_sm, max_scale=200)

vol_mid = (uvol_p[:-1] + uvol_p[1:]) / 2
print(f"  AMPD: {len(peaks_ampd)} 峰 (总 {len(deriv_sm)} 点)")
if len(peaks_ampd):
    pv = deriv_sm[peaks_ampd]
    top3 = np.argsort(np.abs(pv))[-3:][::-1]
    print("  Top-3 (|dV/dt|):")
    for idx in top3:
        print(f"    vol={vol_mid[peaks_ampd[idx]]:.4f} mL  dV/dt={pv[idx]:+.2f}")

# ═══════════════════════════════════════════════════════════════════════
#  3. Spectral cross-entropy analysis
# ═══════════════════════════════════════════════════════════════════════

print("\n[4] 光谱通道分析…")
from DataProcessor.endpoint import _local_maxima


def cross_entropy(arr):
    p = arr[1:].astype(np.float64)
    q = arr[:-1].astype(np.float64)
    p /= p.sum(axis=1, keepdims=True) + 1e-12
    q /= q.sum(axis=1, keepdims=True) + 1e-12
    return -np.sum(p * np.log(np.maximum(q, 1e-12)), axis=1)

ce_8 = cross_entropy(mspec_interp)
ce_721 = cross_entropy(spectra_721)

for label, ce in [("8ch-raw", ce_8), ("721-pt", ce_721)]:
    w = min(15, len(ce) if len(ce) % 2 else len(ce) - 1)
    ce_sm = savgol_filter(ce, window=w, order=2) if w >= 5 else ce
    lm = _local_maxima(ce_sm)
    print(f"  {label}: CE∈[{ce.min():.6e}, {ce.max():.6e}], {len(lm)} 局部极大值")
    if len(lm):
        vals = ce_sm[lm]
        top = np.argsort(vals)[-3:][::-1]
        for idx in top:
            print(f"    vol={vol_mid[lm[idx]]:.4f} mL  CE={vals[idx]:.6e}")

# ═══════════════════════════════════════════════════════════════════════
#  4. Run EndpointDetector with monkey-patched AMPD
# ═══════════════════════════════════════════════════════════════════════

print("\n[5] 运行 EndpointDetector (patched AMPD)…")
import DataProcessor.endpoint as ep_mod

ep_mod._ampd = lambda s, mx=None: ampd_limited(s, mx or 200)

from DataProcessor.endpoint import EndpointDetector

det = EndpointDetector(flow_rate=FLOW_RATE, potential_window=300.0,
                       spectral_window=300.0, max_potential_points=6000,
                       max_spectral_frames=500, consensus_threshold=1.0)

stride = 5
for i in range(0, N, stride):
    det.feed_potential(float(t[i]), float(mpot[i]))
    det.feed_spectrum(float(t[i]), spectra_721[i])

result = det.detect()

# ═══════════════════════════════════════════════════════════════════════
#  5. Visualization
# ═══════════════════════════════════════════════════════════════════════

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt

fig, axes = plt.subplots(3, 1, figsize=(14, 14), sharex=True)
fig.suptitle("TController 滴定终点检测算法验证", fontsize=15, fontweight="bold")

ax = axes[0]
ax.plot(uvol_p, mpot, "b-", lw=0.6, alpha=0.7, label="电位 (LSB)")
ax.set_ylabel("电位 (LSB)")
ax.set_title("(a) 电位 — 体积曲线")
ax.grid(True, alpha=0.3)
ax.legend(loc="upper right")

ax = axes[1]
ax.plot(vol_mid, deriv_sm, "g-", lw=1, label="dV/dt (savgol w=15)")
if len(peaks_ampd):
    ax.scatter(vol_mid[peaks_ampd], deriv_sm[peaks_ampd], c="red", s=20, zorder=5,
               label=f"AMPD 峰 ({len(peaks_ampd)})", alpha=0.6)
    for idx in top3:
        ax.axvline(vol_mid[peaks_ampd[idx]], color="red", ls="--", lw=1, alpha=0.4)
ax.axhline(0, color="gray", lw=0.5)
ax.set_ylabel("dV/dt (LSB/s)")
ax.set_title("(b) 电位一阶导数 + AMPD (max_scale=200, window=15)")
ax.grid(True, alpha=0.3)
ax.legend(loc="upper right")

ax = axes[2]
ce_label = "8ch 交叉熵 (savgol w=15)"
ax.plot(vol_mid, savgol_filter(ce_8, window=15, order=2), "m-", lw=1, label=ce_label)
lm_ce8 = _local_maxima(savgol_filter(ce_8, window=15, order=2))
if len(lm_ce8):
    ax.scatter(vol_mid[lm_ce8], savgol_filter(ce_8, window=15, order=2)[lm_ce8],
               c="red", s=20, zorder=5, label=f"局部极大值 ({len(lm_ce8)})", alpha=0.6)
ax.set_ylabel("交叉熵")
ax.set_title("(c) 光谱 8通道交叉熵")
ax.set_xlabel("体积 (mL)")
ax.grid(True, alpha=0.3)
ax.legend(loc="upper right")

# summary
lines = ["滴定终点检测结果", "─" * 32, ""]
if result:
    lines.append(f"终点体积: {result['volume']:.4f} mL")
    lines.append(f"置信度:   {result['confidence']}")
    lines.append(f"方法:     {result['method']}")
    if result.get("potential"):
        lines.append(f"电位峰:   {result['potential']['volume']:.4f} mL")
    if result.get("spectral"):
        lines.append(f"光谱峰:   {result['spectral']['volume']:.4f} mL")
lines.append("")
lines.append("AMPD max_scale=200")
lines.append(f"数据 {N}步 | 流速 {FLOW_RATE:.6f}")

bbox = {"boxstyle": "round,pad=0.5", "fc": "lightyellow", "alpha": 0.9}
fig.text(0.70, 0.92, "\n".join(lines), fontfamily="monospace", fontsize=9,
         va="top", bbox=bbox, transform=fig.transFigure)

plt.tight_layout(rect=[0, 0, 1, 0.92])
out = OUT_DIR / "validation.png"
plt.savefig(str(out), dpi=150, bbox_inches="tight")
plt.close()
print(f"\n可视化: {out}")

print("\n" + "=" * 60)
print("最终检测结果")
print("=" * 60)
if result:
    icons = {"high": "✅", "medium": "⚠️", "low": "❌"}
    print(f"  {icons.get(result['confidence'], '❓')} 终点: {result['volume']:.4f} mL")
    print(f"    置信度: {result['confidence']}  |  方法: {result['method']}")
    if result.get("warning"): print(f"    警告: {result['warning']}")
    if result.get("potential"):
        p = result["potential"]
        print(f"    电位: {p['volume']:.4f} mL ({p['peak_count']} peaks)")
    if result.get("spectral"):
        s = result["spectral"]
        print(f"    光谱: {s['volume']:.4f} mL")
else:
    print("  ❌ 未检测到终点")
