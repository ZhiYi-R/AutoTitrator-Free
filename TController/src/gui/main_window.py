"""滴定控制主窗口（ttkbootstrap）— 控制台式命令栏 + 工作流相位 + i18n。"""

from __future__ import annotations

import os
import time
import tkinter as tk
from datetime import datetime
from enum import Enum
from tkinter import messagebox
from typing import Literal

import numpy as np
import openpyxl
import ttkbootstrap as ttk
from ttkbootstrap.dialogs import Messagebox

from Communication import ProtocolHandler
from DataProcessor import PUMP_SLOPE, EndpointDetector, steps_from_volume
from DataProcessor import reconstruct as _reconstruct
from DataProcessor._path import CALIBRE_PATH
from gui import i18n, themes
from gui.calibration_tab import CalibrationTab
from gui.maintenance_tab import MaintenanceTab
from gui.potential_widget import PotentialWidget
from gui.results_panel import ResultsPanel
from gui.settings import load_settings, save_settings
from gui.spectrum_widget import SpectrumWidget
from gui.themes import UI_FONT
from gui.widgets import Card, MessageBar, PhaseStepper, StatusDot, TGauge, Tooltip


class TitrationState(Enum):
    IDLE = "idle"
    INJECTING = "injecting"
    TITRATING = "titrating"
    DEGREE_1 = "degree1"
    TITRATING_2 = "titrating2"
    DONE = "done"
    ERROR = "error"


# 状态 → 芯片样式
_CHIP_STYLE: dict[TitrationState, str] = {
    TitrationState.IDLE: "ChipIdle",
    TitrationState.INJECTING: "ChipRun",
    TitrationState.TITRATING: "ChipRun",
    TitrationState.TITRATING_2: "ChipRun",
    TitrationState.DEGREE_1: "ChipWarn",
    TitrationState.DONE: "ChipOk",
    TitrationState.ERROR: "ChipErr",
}

# 状态 → 相位步进索引（ERROR 保持当前相位，仅标红）
_PHASE_INDEX: dict[TitrationState, int] = {
    TitrationState.IDLE: 0,
    TitrationState.INJECTING: 1,
    TitrationState.TITRATING: 2,
    TitrationState.DEGREE_1: 3,
    TitrationState.TITRATING_2: 3,
    TitrationState.DONE: 4,
    TitrationState.ERROR: 0,
}

_PHASES = [
    ("ready", "phases.ready"),
    ("injecting", "phases.injecting"),
    ("titrating", "phases.titrating"),
    ("detecting", "phases.detecting"),
    ("done", "phases.done"),
]


# ======================================================================


class MainWindow(ttk.Frame):
    """滴定控制主窗口。

    工作流:
      空闲 → [开始] → 进样泵 MaxCount → 滴定泵 FreeRun
      → 终点检测 T=1 → 继续 FreeRun → T=2 → 停止
    """

    def __init__(
        self,
        parent: tk.Misc,
        theme_mode: str = "system",
        **kwargs,
    ) -> None:
        super().__init__(parent, **kwargs)

        # ---- 串口通信 ----
        self._com = ProtocolHandler()
        self._com.on("connected", self._on_connected)
        self._com.on("disconnected", self._on_disconnected)
        self._com.on("error", self._on_com_error)
        self._com.on("spectral", self._on_spectral_data)
        self._com.on("adc", self._on_adc_data)
        self._com.on("pump_done", self._on_pump_done)
        self._com.on("pump1_progress", self._on_pump1_progress)
        self._com.on("pump2_progress", self._on_pump2_progress)
        self._com.on("nak", self._on_nak)

        self._pump1_volume: float = 0.0  # 泵 1 体积（mL）
        self._pump2_volume: float = 0.0  # 泵 2 体积（mL）

        # ---- 滴定检测 ----
        self._detector = EndpointDetector()
        self._state = TitrationState.IDLE
        self._phase_idx = 0
        self._endpoint_volume: float | None = None
        self._t0: float = 0.0  # 连接时刻
        self._t1_time: float = 0.0
        self._recon_wls: np.ndarray | None = None
        self._adc_buffer: list[int] = []
        self._adc_counter = 0

        # ---- 数据记录 ----
        self._recording = False
        self._rec_spectral: list[tuple[float, list[int]]] = []
        self._rec_recon: list[tuple[float, np.ndarray]] = []
        self._rec_recon_wls: np.ndarray | None = None
        self._rec_potential: list[tuple[float, float, float, float]] = []
        self._rec_raw_adc: list[tuple[float, int, float]] = []
        self._rec_ewma_v: float | None = None

        # ---- UI ----
        self._theme_mode = theme_mode if theme_mode in themes.MODES else "system"
        self._connected = False
        self._rec_var = tk.BooleanVar(value=True)
        self._group_captions: list[tuple[ttk.Label, str]] = []
        self._build_toolbar()
        # 状态栏先于中央区打包（side=bottom），确保窗口高度不足时不被挤压
        self._build_status()
        self._build_central()
        self._load_electrodes()
        self._bind_shortcuts()
        i18n.subscribe(self._apply_i18n)
        themes.subscribe(self._apply_theme)

        # 初始空状态提示
        self._spectrum_widget.set_overlay("overlay.disconnected")
        self._potential_widget.set_overlay("overlay.disconnected")

        # ---- 定时器（root.after 递归调度）----
        self._running = True
        self._heartbeat_active = False
        self._schedule_after(80, self._refresh_plots)
        self._schedule_after(2000, self._scan_ports)
        self._scan_ports()
        self._schedule_after(500, self._run_detection)
        # 心跳定时器由连接/断开控制

    def _schedule_after(self, ms: int, callback) -> None:
        """安全的 after 调度：窗口关闭后不再触发。"""

        def _wrapper():
            if not self._running:
                return
            callback()
            self._schedule_after(ms, callback)

        self.after(ms, _wrapper)

    def _schedule_heartbeat(self) -> None:
        def _heartbeat():
            if not self._running or not self._heartbeat_active:
                return
            self._send_heartbeat()
            self.after(1000, _heartbeat)

        if self._heartbeat_active:
            self.after(1000, _heartbeat)

    # ================================================================
    #  UI 构建
    # ================================================================

    def _group(
        self,
        parent: tk.Misc,
        caption_key: str,
        side: Literal["left", "right", "top", "bottom"] = "left",
        padx: tuple[int, int] = (0, 8),
    ) -> ttk.Frame:
        """带标题的工具栏分组卡片（已入包），返回控件行容器。"""
        card = Card(parent, tone="surface")
        card.pack(side=side, padx=padx)
        body = ttk.Frame(card, style="Toolbar.TFrame", padding=(10, 4))
        body.pack(fill="both", expand=True)
        cap = ttk.Label(body, text=i18n.tr(caption_key), style="GroupCaption.TLabel")
        cap.pack(anchor="w")
        row = ttk.Frame(body, style="Toolbar.TFrame")
        row.pack(anchor="w", pady=(2, 3))
        self._group_captions.append((cap, caption_key))
        return row

    def _build_toolbar(self) -> None:
        tb = ttk.Frame(self, style="Toolbar.TFrame", padding=(8, 6))
        tb.pack(fill="x")

        # ---- 连接组 ----
        conn_row = self._group(tb, "toolbar.group_conn")

        self._status_dot = StatusDot(conn_row)
        self._status_dot.pack(side="left", padx=(0, 4))

        self._port_label = ttk.Label(conn_row, text=i18n.tr("toolbar.port"), style="GroupLabel.TLabel")
        self._port_label.pack(side="left", padx=(0, 4))
        self._port_cb = ttk.Combobox(conn_row, state="readonly", width=16)
        self._port_cb.pack(side="left")
        self._port_data: dict[str, str] = {}  # 显示文本 → 设备路径

        self._more_btn = ttk.Menubutton(
            conn_row, text=i18n.tr("toolbar.more"), width=5, bootstyle="outline"
        )
        self._more_menu = tk.Menu(self._more_btn, tearoff=0)
        self._more_btn["menu"] = self._more_menu
        self._more_btn.pack(side="left", padx=4)

        self._baud_label = ttk.Label(conn_row, text=i18n.tr("toolbar.baud"), style="GroupLabel.TLabel")
        self._baud_label.pack(side="left", padx=(8, 4))
        self._baud_cb = ttk.Combobox(
            conn_row,
            values=["9600", "19200", "38400", "57600", "115200", "230400"],
            state="readonly",
            width=6,
        )
        self._baud_cb.set(str(load_settings()["baud"]))
        self._baud_cb.pack(side="left")

        self._conn_btn = ttk.Button(
            conn_row,
            text=i18n.tr("toolbar.connect"),
            bootstyle="primary",
            command=self._toggle_connect,
        )
        self._conn_btn.pack(side="left", padx=(8, 0))

        # ---- 滴定控制组 ----
        run_row = self._group(tb, "toolbar.group_run")

        self._vol_label = ttk.Label(
            run_row, text=i18n.tr("toolbar.sample_volume"), style="GroupLabel.TLabel"
        )
        self._vol_label.pack(side="left", padx=(0, 4))
        self._vol_spin = ttk.Spinbox(
            run_row, from_=1.0, to=999.0, increment=0.1, format="%.1f", width=5
        )
        self._vol_spin.set(5.0)
        self._vol_spin.pack(side="left")

        self._start_btn = ttk.Button(
            run_row,
            text=i18n.tr("toolbar.start"),
            bootstyle="success",
            padding=(10, 4),
            command=self._start_titration,
        )
        self._start_btn.pack(side="left", padx=(10, 4))
        self._start_btn.state(["disabled"])
        Tooltip(self._start_btn, lambda: i18n.tr("tooltip.start"))

        self._manual_stop_btn = ttk.Button(
            run_row,
            text=i18n.tr("toolbar.stop"),
            bootstyle="outline",
            command=self._manual_stop,
        )
        self._manual_stop_btn.pack(side="left", padx=(0, 6))
        self._manual_stop_btn.state(["disabled"])

        # ---- 设备组 ----
        dev_row = self._group(tb, "toolbar.group_dev", padx=(0, 0))

        self._reset_btn = ttk.Button(
            dev_row,
            text=i18n.tr("toolbar.reset_mcu"),
            bootstyle="outline",
            command=self._reset_mcu,
        )
        self._reset_btn.pack(side="left")

        # ---- 急停区（右端独立危险区）----
        estop_zone = Card(tb, tone="danger")
        estop_zone.pack(side="right", padx=(8, 0))
        estop_body = ttk.Frame(estop_zone, style="EstopZoneBody.TFrame", padding=(12, 4))
        estop_body.pack(fill="both", expand=True)
        self._estop_cap = ttk.Label(
            estop_body, text=i18n.tr("toolbar.emergency_stop"), style="EstopZone.TLabel"
        )
        self._estop_cap.pack(anchor="w")
        self._stop_btn = ttk.Button(
            estop_body,
            text=i18n.tr("toolbar.emergency_stop"),
            bootstyle="danger",
            padding=(14, 6),
            command=self._emergency_stop,
        )
        self._stop_btn.pack(pady=(1, 3))
        self._stop_btn.state(["disabled"])
        Tooltip(self._stop_btn, lambda: i18n.tr("tooltip.estop"))

        # ---- 显示组（语言 / 主题）----
        view_row = self._group(tb, "toolbar.group_view", side="right", padx=(8, 0))

        self._lang_cb = ttk.Combobox(view_row, state="readonly", width=8)
        self._lang_cb.pack(side="left", padx=(0, 4))
        self._rebuild_lang_combo()
        self._lang_cb.bind("<<ComboboxSelected>>", self._on_lang_changed)

        self._theme_cb = ttk.Combobox(view_row, state="readonly", width=8)
        self._theme_cb.pack(side="left")
        self._rebuild_theme_combo()
        self._theme_cb.bind("<<ComboboxSelected>>", self._on_theme_changed)

    def _rebuild_theme_combo(self) -> None:
        """按当前语言重建主题下拉框（保持选中项）。"""
        keys = {"theme.light": "light", "theme.dark": "dark", "theme.system": "system"}
        self._theme_display_to_mode = {i18n.tr(k): m for k, m in keys.items()}
        self._theme_cb["values"] = list(self._theme_display_to_mode.keys())
        for disp, mode in self._theme_display_to_mode.items():
            if mode == self._theme_mode:
                self._theme_cb.set(disp)
                break

    def _rebuild_lang_combo(self) -> None:
        """重建语言下拉框（显示各语言自称）。"""
        self._lang_display_to_code = {name: code for code, name in i18n.LANGS.items()}
        self._lang_cb["values"] = list(self._lang_display_to_code.keys())
        self._lang_cb.set(i18n.LANGS[i18n.current_language()])

    def _build_central(self) -> None:
        self._main_tabs = ttk.Notebook(self)
        self._main_tabs.pack(fill="both", expand=True, padx=4, pady=(2, 0))

        # ---- 滴定标签页 ----
        titrate_tab = ttk.Frame(self._main_tabs)

        # 工作流相位 + 滴定度规
        top = ttk.Frame(titrate_tab)
        top.pack(fill="x", padx=8)
        self._stepper = PhaseStepper(top, _PHASES)
        self._stepper.pack(side="left", fill="x", expand=True)
        ttk.Separator(top, orient="vertical").pack(side="left", fill="y", padx=10, pady=8)
        self._tgauge = TGauge(top, width=340)
        self._tgauge.pack(side="right", padx=(6, 4))

        paned = tk.PanedWindow(
            titrate_tab, orient="horizontal", sashwidth=4
        )
        paned.pack(fill="both", expand=True)

        left = ttk.Frame(paned)
        left.pack(fill="both", expand=True)

        self._spectrum_widget = SpectrumWidget(left)
        self._spectrum_widget.pack(fill="both", expand=True, pady=(0, 2))

        self._potential_widget = PotentialWidget(left)
        self._potential_widget.pack(fill="both", expand=True)

        paned.add(left, minsize=400, stretch="always")

        self._results_panel = ResultsPanel(
            paned,
            record_var=self._rec_var,
            record_command=self._on_recording_toggled,
        )
        paned.add(self._results_panel, minsize=260, stretch="never")

        self._main_tabs.add(titrate_tab, text=i18n.tr("tabs.titration"))

        # ---- 校准标签页 ----
        self._calib_tab = CalibrationTab(self._com, parent=self._main_tabs)
        self._main_tabs.add(self._calib_tab, text=i18n.tr("tabs.calibration"))

        self._maintenance_tab = MaintenanceTab(
            self._com, parent=self._main_tabs
        )
        self._main_tabs.add(self._maintenance_tab, text=i18n.tr("tabs.maintenance"))
        self._maintenance_tab.set_connected(False)

    def _build_status(self) -> None:
        sb = ttk.Frame(self, style="Statusbar.TFrame", padding=(10, 5))
        sb.pack(fill="x", side="bottom")

        self._chip = ttk.Label(sb, style="ChipIdle.TLabel", text=i18n.tr("states.idle"))
        self._chip.pack(side="left")

        self._msg = MessageBar(sb)
        self._msg.pack(side="left", fill="x", expand=True, padx=(12, 0))

        self._activity = ttk.Label(sb, text="", style="Status.TLabel", font=(UI_FONT, 8))
        self._activity.pack(side="right", padx=(12, 10))

        self._conn_dot = StatusDot(sb)
        self._conn_dot.pack(side="right", padx=(0, 4))
        self._conn_label = ttk.Label(sb, text="", style="Conn.TLabel")
        self._conn_label.pack(side="right")
        self._update_conn_label()

    def _bind_shortcuts(self) -> None:
        """全局快捷键：Esc 急停，F5 开始滴定。"""
        top = self.winfo_toplevel()
        top.bind("<Escape>", lambda _e: self._shortcut_estop())
        top.bind("<F5>", lambda _e: self._shortcut_start())

    def _shortcut_estop(self) -> None:
        if self._connected:
            self._emergency_stop()

    def _shortcut_start(self) -> None:
        if "disabled" not in self._start_btn.state():
            self._start_titration()

    # ---- i18n / 主题刷新 ----

    def _apply_i18n(self) -> None:
        self.winfo_toplevel().title(i18n.tr("app.title"))
        for cap, key in self._group_captions:
            cap.config(text=i18n.tr(key))
        self._port_label.config(text=i18n.tr("toolbar.port"))
        self._baud_label.config(text=i18n.tr("toolbar.baud"))
        self._more_btn.config(text=i18n.tr("toolbar.more"))
        self._conn_btn.config(
            text=i18n.tr("toolbar.disconnect" if self._connected else "toolbar.connect")
        )
        self._vol_label.config(text=i18n.tr("toolbar.sample_volume"))
        self._start_btn.config(text=i18n.tr("toolbar.start"))
        self._manual_stop_btn.config(text=i18n.tr("toolbar.stop"))
        self._stop_btn.config(text=i18n.tr("toolbar.emergency_stop"))
        self._estop_cap.config(text=i18n.tr("toolbar.emergency_stop"))
        self._reset_btn.config(text=i18n.tr("toolbar.reset_mcu"))
        self._main_tabs.tab(0, text=i18n.tr("tabs.titration"))
        self._main_tabs.tab(1, text=i18n.tr("tabs.calibration"))
        self._main_tabs.tab(2, text=i18n.tr("tabs.maintenance"))
        # 状态芯片与连接标签（状态枚举名来自 i18n）
        self._set_chip(i18n.tr(f"states.{self._state.value}"), _CHIP_STYLE[self._state])
        self._update_conn_label()
        # 下拉框选项随语言重建
        self._rebuild_theme_combo()
        self._rebuild_lang_combo()

    def _apply_theme(self) -> None:
        self._update_conn_label()
        themes.set_native_titlebar(themes.current_key() == "dark", self.winfo_toplevel())

    def _set_chip(self, text: str, style: str) -> None:
        self._chip.config(text=text, style=f"{style}.TLabel")

    def _update_conn_label(self) -> None:
        t = themes.current_tokens()
        if self._connected:
            disp = self._port_cb.get()
            port = self._port_data.get(disp, disp)
            self._conn_label.config(
                text=port if port else i18n.tr("conn.connected"), foreground=t.fg
            )
            self._conn_dot.set_state("ok")
            self._status_dot.set_state("ok")
        else:
            self._conn_label.config(text=i18n.tr("conn.disconnected"), foreground=t.fg_muted)
            self._conn_dot.set_state("off")
            self._status_dot.set_state("off")

    # ================================================================
    #  串口连接
    # ================================================================

    def _scan_ports(self) -> None:
        import serial.tools.list_ports

        current = self._port_cb.get()
        self._port_data.clear()
        values = []
        for p in serial.tools.list_ports.comports():
            if p.vid is not None:
                desc = p.description or ""
                label = f"{p.device} ({desc})" if desc else p.device
                values.append(label)
                self._port_data[label] = p.device

        self._port_cb["values"] = values
        if values:
            if current in values:
                self._port_cb.set(current)
            else:
                # 优先恢复上次连接的端口
                last = load_settings()["last_port"]
                restored = next(
                    (lbl for lbl, dev in self._port_data.items() if dev == last), None
                )
                self._port_cb.set(restored or values[0])

        # 无 vid 的端口放入"更多"菜单
        self._more_menu.delete(0, "end")
        bare = [p.device for p in serial.tools.list_ports.comports() if p.vid is None]
        if bare:
            self._more_btn.state(["!disabled"])
            for dev in bare:
                self._more_menu.add_command(
                    label=dev,
                    command=lambda d=dev: self._on_bare_port(d),
                )
        else:
            self._more_btn.state(["disabled"])

    def _on_bare_port(self, dev: str) -> None:
        label = dev
        self._port_data[label] = dev
        values = list(self._port_cb["values"])
        if label not in values:
            values.append(label)
            self._port_cb["values"] = values
        self._port_cb.set(label)

    def _toggle_connect(self) -> None:
        if self._com.is_open:
            self._com.send_abort(0xFF)
            self._com.disconnect()
        else:
            disp = self._port_cb.get()
            port = self._port_data.get(disp, disp)
            if not port:
                messagebox.showwarning(i18n.tr("conn.hint"), i18n.tr("conn.select_port"))
                return
            self._com.reconfigure(port, int(self._baud_cb.get()))
            self._com.connect()

    def _on_connected(self, _data: object = None) -> None:
        self._connected = True
        self._conn_btn.config(
            text=i18n.tr("toolbar.disconnect"), bootstyle="outline"
        )
        self._t0 = time.monotonic()
        self._update_conn_label()
        self._start_btn.state(["!disabled"])
        self._stop_btn.state(["!disabled"])  # 急停：连接后始终可用
        self._com.enable_watchdog()
        self._heartbeat_active = True
        self._schedule_heartbeat()
        self._maintenance_tab.set_connected(True)
        # 记住端口，恢复空状态
        disp = self._port_cb.get()
        port = self._port_data.get(disp, disp)
        save_settings(last_port=port, baud=int(self._baud_cb.get()))
        self._spectrum_widget.set_overlay("overlay.waiting")
        self._potential_widget.set_overlay("overlay.waiting")
        self._msg.show("success", i18n.tr("msg.connected", port=port), sticky=True)

    def _on_disconnected(self, _data: object = None) -> None:
        self._connected = False
        self._conn_btn.config(text=i18n.tr("toolbar.connect"), bootstyle="primary")
        self._update_conn_label()
        self._start_btn.state(["disabled"])
        self._stop_btn.state(["disabled"])
        self._set_state(TitrationState.IDLE)
        self._heartbeat_active = False
        self._maintenance_tab.set_connected(False)
        self._spectrum_widget.set_overlay("overlay.disconnected")
        self._potential_widget.set_overlay("overlay.disconnected")
        self._tgauge.set_value(None)
        self._msg.show("warn", i18n.tr("msg.disconnected"), sticky=True)

    def _on_com_error(self, msg: str) -> None:
        self._msg.show("error", i18n.tr("status.error_fmt", msg=msg), sticky=True)
        self._set_state(TitrationState.ERROR)

    def _on_nak(self, _data: object = None) -> None:
        self._activity.config(text=i18n.tr("status.nak"))

    def _send_heartbeat(self) -> None:
        if self._com.is_open:
            self._com.send_heartbeat()

    # ================================================================
    #  数据回调
    # ================================================================

    def _on_spectral_data(self, vals: list[int]) -> None:
        t = time.monotonic() - self._t0
        # 始终重建并显示
        if self._recon_wls is None:
            try:
                self._recon_wls, _ = _reconstruct(vals)
            except Exception:
                pass
        if self._recon_wls is not None:
            try:
                _, spec = _reconstruct(vals)
                self._spectrum_widget.update_spectrum(self._recon_wls, spec)
                if self._recording:
                    self._rec_spectral.append((t, list(vals)))
                    if self._rec_recon_wls is None:
                        self._rec_recon_wls = self._recon_wls.copy()
                    self._rec_recon.append((t, spec.copy()))
            except Exception:
                pass
        # 仅在滴定期间馈入检测器
        if self._state in (
            TitrationState.TITRATING,
            TitrationState.TITRATING_2,
            TitrationState.DEGREE_1,
        ):
            self._detector.feed_spectrum(
                self._pump2_volume, np.array(vals, dtype=np.float64)
            )

    def _on_adc_data(self, data: tuple) -> None:
        raw, pump2_pos = data
        t = time.monotonic() - self._t0
        v = raw * 3.3 / 65535 - 1.1
        # 更新 Pump2 体积（基于固件实际步数）
        old_vol = self._pump2_volume
        self._pump2_volume = PUMP_SLOPE * pump2_pos
        self._update_gauge()
        # 泵体积变化时更新活动区
        if self._state in (
            TitrationState.TITRATING,
            TitrationState.TITRATING_2,
            TitrationState.DEGREE_1,
        ):
            diff = self._pump2_volume - old_vol
            if diff > 0.001:
                self._activity.config(
                    text=i18n.tr(
                        "status.pump2_vol",
                        vol=f"{self._pump2_volume:.4f}",
                        diff=f"{diff:.4f}",
                    )
                )
        # 缓存用于泵 report 间平均
        self._adc_buffer.append(raw)
        self._results_panel.set_current_voltage(v)
        # 记录原始 ADC 数据
        if self._recording:
            self._rec_raw_adc.append((t, raw, self._pump2_volume))
        # 每次均馈入检测器（保持检测精度）
        if self._state in (
            TitrationState.TITRATING,
            TitrationState.TITRATING_2,
            TitrationState.DEGREE_1,
        ):
            vol = self._pump2_volume
            self._detector.feed_potential(vol, t, v)
        self._adc_counter += 1

    def _on_pump1_progress(self, pos: int) -> None:
        self._pump1_volume = PUMP_SLOPE * pos
        self._results_panel.update_inject_progress(pos, volume=self._pump1_volume)
        self._flush_adc_buffer()

    def _on_pump2_progress(self, pos: int) -> None:
        self._pump2_volume = PUMP_SLOPE * pos
        self._update_gauge()
        self._flush_adc_buffer()

    def _flush_adc_buffer(self) -> None:
        """将缓存内 ADC 读数平均后更新到电位曲线。"""
        if not self._adc_buffer:
            return
        t = time.monotonic() - self._t0
        raw_avg = round(sum(self._adc_buffer) / len(self._adc_buffer))
        v = raw_avg * 3.3 / 65535 - 1.1
        vol = (
            self._pump1_volume
            if self._state == TitrationState.INJECTING
            else self._pump2_volume
        )
        self._potential_widget.append(t, raw_avg, volume=vol)
        # 记录电位数据点（含 EWMA 平滑）
        if self._recording:
            if self._rec_ewma_v is None:
                self._rec_ewma_v = v
            else:
                self._rec_ewma_v = 0.15 * v + 0.85 * self._rec_ewma_v
            self._rec_potential.append((t, v, self._rec_ewma_v, vol))
        self._adc_buffer.clear()

    def _on_pump_done(self, data: tuple) -> None:
        pump_id, position = data
        self._activity.config(text=i18n.tr("status.pump_done", id=pump_id, pos=position))
        if pump_id == 1 and self._state == TitrationState.INJECTING:
            # 进样完成 -> 隐藏进度条，清空进样数据，启动滴定泵
            self._results_panel.hide_inject_progress()
            self._potential_widget.reset()
            self._potential_widget.set_titrating(True)
            self._set_state(TitrationState.TITRATING)
            self._com.send_frerun(2)

    # ================================================================
    #  滴定控制
    # ================================================================

    def _update_gauge(self) -> None:
        if self._endpoint_volume and self._state in (
            TitrationState.TITRATING,
            TitrationState.DEGREE_1,
            TitrationState.TITRATING_2,
            TitrationState.DONE,
        ):
            self._tgauge.set_value(self._pump2_volume / self._endpoint_volume)
        else:
            self._tgauge.set_value(None)

    def _set_state(self, state: TitrationState) -> None:
        self._state = state
        self._set_chip(i18n.tr(f"states.{state.value}"), _CHIP_STYLE[state])
        if state is TitrationState.ERROR:
            self._stepper.set_phase(self._phase_idx, error=True)
        else:
            self._phase_idx = _PHASE_INDEX[state]
            self._stepper.set_phase(
                self._phase_idx, done_all=state is TitrationState.DONE
            )

    def _start_titration(self) -> None:
        if not self._com.is_open:
            messagebox.showwarning(i18n.tr("conn.hint"), i18n.tr("conn.connect_first"))
            return

        # 复位检测器与所有数据
        self._detector.reset()
        self._endpoint_volume = None
        self._t1_time = 0.0
        self._potential_widget.set_titrating(False)
        self._potential_widget.reset()
        self._results_panel.reset_endpoint()
        self._tgauge.set_value(None)
        self._adc_counter = 0
        # AMPD 在 T=2 统一执行，无需中间标记
        self._pump1_volume = 0.0
        self._pump2_volume = 0.0
        try:
            vol_ml = float(self._vol_spin.get())
        except (ValueError, tk.TclError):
            vol_ml = 5.0
        self._msg.show("info", i18n.tr("status.injecting", vol=f"{vol_ml:.1f}"), sticky=True)
        self._results_panel.set_sample_volume(vol_ml)
        # 进样体积 → MaxCount 步数（标定公式换算）
        steps = steps_from_volume(vol_ml)
        self._results_panel.show_inject_progress(steps, target_vol=vol_ml)
        self._start_btn.state(["disabled"])
        self._com.send_maxcount(1, steps)
        self._set_state(TitrationState.INJECTING)
        self._potential_widget.set_titrating(True)
        self._manual_stop_btn.state(["!disabled"])

    def _emergency_stop(self) -> None:
        self._com.send_reset()
        self._set_state(TitrationState.IDLE)
        self._start_btn.state(["!disabled"])
        self._manual_stop_btn.state(["disabled"])
        self._msg.show("warn", i18n.tr("status.emergency_stopped"), sticky=True)
        self._potential_widget.set_titrating(False)
        self._potential_widget.reset()
        self._results_panel.reset_endpoint()
        self._tgauge.set_value(None)

    def _manual_stop(self) -> None:
        """手动停止滴定：停泵、精修终点、保存数据，不复位 MCU。"""
        self._manual_stop_btn.state(["disabled"])

        # 停止滴定泵
        self._com.send_frestop(2)

        # 如果已检测到 T=1，用 AMPD 精修终点
        if self._endpoint_volume:
            ref = self._detector.refine_with_ampd()
            if ref is not None:
                self._endpoint_volume = ref
                self._results_panel.set_endpoint(ref)
                self._potential_widget.set_endpoint(ref)

            self._potential_widget.set_titrating(False)
            self._set_state(TitrationState.DONE)
            self._set_chip(
                i18n.tr("status.manual_stopped_at", vol=f"{self._endpoint_volume:.4f}"),
                "ChipWarn",
            )
            self._msg.show(
                "warn",
                i18n.tr("status.manual_stopped_at", vol=f"{self._endpoint_volume:.4f}"),
                sticky=True,
            )

            # 保存数据
            result = self._detector.detect()
            if result is not None:
                self._on_titration_complete(result)
        else:
            self._potential_widget.set_titrating(False)
            self._set_state(TitrationState.DONE)
            self._set_chip(i18n.tr("status.manual_stopped"), "ChipWarn")
            self._msg.show("warn", i18n.tr("status.manual_stopped"), sticky=True)

        self._start_btn.state(["!disabled"])

    def _reset_mcu(self) -> None:
        """复位 MCU 并重置上位机状态（需确认）。"""
        answer = Messagebox.yesno(
            i18n.tr("confirm.reset_msg"),
            i18n.tr("confirm.reset_title"),
            parent=self.winfo_toplevel(),
            alert=True,
        )
        if answer != "Yes":
            return
        self._com.send_reset()
        self._detector.reset()
        self._set_state(TitrationState.IDLE)
        self._start_btn.state(["!disabled"])
        self._manual_stop_btn.state(["disabled"])
        self._potential_widget.set_titrating(False)
        self._potential_widget.reset()
        self._results_panel.reset_endpoint()
        self._endpoint_volume = None
        self._t1_time = 0.0
        self._tgauge.set_value(None)
        self._msg.show("info", i18n.tr("status.reset_done"), sticky=True)

    # ================================================================
    #  定时任务
    # ================================================================

    def _refresh_plots(self) -> None:
        # 排空通信事件队列
        self._com.poll()
        self._potential_widget.refresh()

    def _run_detection(self) -> None:
        if self._state not in (
            TitrationState.TITRATING,
            TitrationState.TITRATING_2,
            TitrationState.DEGREE_1,
        ):
            return

        result = self._detector.detect()
        if result is None:
            return

        vol = result["volume"]
        self._activity.config(
            text=i18n.tr("status.detect", vol=f"{vol:.3f}", conf=result["confidence"])
        )

        if self._state == TitrationState.TITRATING:
            # 首次到达终点 T=1
            self._endpoint_volume = vol
            self._potential_widget.set_endpoint(vol)
            self._results_panel.set_endpoint(vol)
            self._set_state(TitrationState.DEGREE_1)
            self._set_chip(i18n.tr("status.endpoint_t1", vol=f"{vol:.3f}"), "ChipWarn")
            self._msg.show(
                "warn", i18n.tr("status.endpoint_t1", vol=f"{vol:.3f}"), sticky=True
            )

        elif self._state in (TitrationState.DEGREE_1, TitrationState.TITRATING_2):
            # 显示当前滴定进度
            if self._endpoint_volume:
                t_val = self._pump2_volume / self._endpoint_volume
                self._set_chip(
                    i18n.tr(
                        "status.progress",
                        ep=f"{self._endpoint_volume:.3f}",
                        t=f"{t_val:.2f}",
                        vol=f"{self._pump2_volume:.3f}",
                    ),
                    "ChipRun",
                )
            # 继续滴定到 T=2（用实际累积体积）
            if (
                self._endpoint_volume
                and self._pump2_volume >= 2.0 * self._endpoint_volume
            ):
                self._com.send_frestop(2)
                # T=2 停泵后用 AMPD 从完整导数曲线精确定位终点
                ref = self._detector.refine_with_ampd()
                if ref is not None:
                    self._endpoint_volume = ref
                    self._results_panel.set_endpoint(ref)
                    self._potential_widget.set_endpoint(ref)
                self._set_state(TitrationState.DONE)
                self._potential_widget.set_titrating(False)
                self._set_chip(
                    i18n.tr("status.endpoint_final", vol=f"{self._endpoint_volume:.4f}"),
                    "ChipOk",
                )
                self._msg.show(
                    "success",
                    i18n.tr("status.endpoint_final", vol=f"{self._endpoint_volume:.4f}"),
                    sticky=True,
                )
                self._start_btn.state(["!disabled"])
                self._manual_stop_btn.state(["disabled"])
                self._on_titration_complete(result)
            else:
                if self._state == TitrationState.DEGREE_1:
                    self._set_state(TitrationState.TITRATING_2)

    # ================================================================
    #  语言 / 主题切换
    # ================================================================

    def _on_lang_changed(self, _event: object = None) -> None:
        code = self._lang_display_to_code.get(self._lang_cb.get())
        if code and code != i18n.current_language():
            save_settings(language=code)
            i18n.set_language(code)

    def _on_theme_changed(self, _event: object = None) -> None:
        mode = self._theme_display_to_mode.get(self._theme_cb.get(), "system")
        self._theme_mode = mode
        save_settings(theme_mode=mode)
        themes.apply_theme(mode, plots=self._theme_plots())

    def _theme_plots(self) -> list:
        plots = [
            self._spectrum_widget,
            self._potential_widget,
        ]
        plots.extend(self._calib_tab.plots)
        return [p for p in plots if p is not None]

    # ================================================================
    #  数据导出
    # ================================================================

    def _on_titration_complete(self, result: dict) -> None:
        """滴定完成时更新 Cx 并自动导出数据。"""
        # 确保 Cx 基于最新终点体积重新计算
        if self._endpoint_volume:
            self._results_panel.set_endpoint(self._endpoint_volume)
        if not self._recording:
            return
        self._export_recording(result)

    def _export_recording(self, result: dict) -> None:
        """将记录的数据写入 ExpResults/ 目录下的 xlsx 文件。"""

        out_dir = os.path.join(os.path.dirname(__file__), "..", "..", "ExpResults")
        os.makedirs(out_dir, exist_ok=True)

        c_std = self._results_panel._spin_value(self._results_panel._c_std)
        ts = datetime.now().strftime("%y-%m-%d-%H-%M-%S")  # noqa: DTZ005 文件名时间戳用本地时间
        filename = f"std_conc_{c_std}_{ts}.xlsx"
        filepath = os.path.join(out_dir, filename)

        wb = openpyxl.Workbook()

        # ---- Sheet 1: Raw Spectrum ----
        ws = wb.active
        ws.title = "Raw Spectrum"
        ws.append(
            [
                "Time (s)",
                "F1(415)", "F2(445)", "F3(480)", "F4(515)",
                "F5(555)", "F6(590)", "F7(630)", "F8(680)",
                "Clear", "NIR(910)",
            ]
        )
        for t, vals in self._rec_spectral:
            ws.append([round(t, 3)] + vals)

        # ---- Sheet 2: Reconstructed Spectrum ----
        if self._rec_recon_wls is not None and self._rec_recon:
            ws2 = wb.create_sheet("Reconstructed Spectrum")
            header = ["Wavelength (nm)"] + [
                f"t={round(t, 3)}s" for t, _ in self._rec_recon
            ]
            ws2.append(header)
            wls = self._rec_recon_wls
            for i in range(len(wls)):
                row = [round(float(wls[i]), 2)]
                for _, spec in self._rec_recon:
                    row.append(round(float(spec[i]), 4))
                ws2.append(row)

        # ---- Sheet 3: Potential (Filtered) ----
        ws3 = wb.create_sheet("Potential")
        ws3.append(
            ["Time (s)", "Raw Voltage (V)", "Filtered Voltage (V)", "Volume (mL)"]
        )
        for t, rv, fv, vol in self._rec_potential:
            ws3.append([round(t, 3), round(rv, 6), round(fv, 6), round(vol, 6)])

        # ---- Sheet 4: Titration Results ----
        ws4 = wb.create_sheet("Titration Results")
        ws4.append(["Parameter", "Value"])
        ws4.append(
            [
                "Endpoint Volume (mL)",
                round(self._endpoint_volume or result.get("volume", 0), 4),
            ]
        )
        ws4.append(["Confidence", result.get("confidence", "")])
        ws4.append(["Method", result.get("method", "")])
        if result.get("potential"):
            ws4.append(
                [
                    "Potential Endpoint (mL)",
                    round(result["potential"].get("volume", 0), 4),
                ]
            )
            ws4.append(["Min dV/dt", result["potential"].get("min_dvdt", "")])
        if result.get("spectral"):
            ws4.append(
                [
                    "Spectral Endpoint (mL)",
                    round(result["spectral"].get("volume", 0), 4),
                ]
            )
            ws4.append(["Max CE", result["spectral"].get("max_ce", "")])
        ws4.append(["C_std (mol/L)", c_std])
        ws4.append(
            ["n_std (标准液)", self._results_panel._spin_value(self._results_panel._n_std)]
        )
        ws4.append(
            [
                "n_analyte (待测液)",
                self._results_panel._spin_value(self._results_panel._n_analyte),
            ]
        )
        ws4.append(["V_sample (mL)", self._results_panel._v_sample_label.cget("text")])
        ws4.append(["Cx (mol/L)", self._results_panel._c_x_label.cget("text")])
        if self._endpoint_volume:
            ws4.append(["Refined Endpoint (mL)", round(self._endpoint_volume, 4)])

        wb.save(filepath)
        self._msg.show("success", i18n.tr("status.exported", file=filename))

    # ================================================================

    def _on_recording_toggled(self) -> None:
        on = self._rec_var.get()
        self._recording = on
        save_settings(record=on)
        if on:
            self._rec_spectral.clear()
            self._rec_recon.clear()
            self._rec_recon_wls = None
            self._rec_potential.clear()
            self._rec_raw_adc.clear()
            self._rec_ewma_v = None
            self._msg.show("info", i18n.tr("status.recording_on"))
        else:
            self._msg.show("info", i18n.tr("status.recording_off"))

    # ---- 电极选择 ----

    def _load_electrodes(self) -> None:
        path = CALIBRE_PATH
        electrodes = []
        if os.path.isfile(path):
            try:
                import numpy as _np

                data = _np.load(path, allow_pickle=True)
                if "n_electrodes" in data:
                    n = int(data["n_electrodes"])
                    names = list(data["names"])
                    for i in range(n):
                        name = names[i]
                        unit = str(data[f"unit_{i}"])
                        slope = float(data[f"slope_{i}"])
                        intercept = float(data[f"intercept_{i}"])
                        electrodes.append((name, slope, intercept, unit))
            except Exception:
                pass
        self._results_panel.set_electrodes(electrodes)
        self._results_panel.on_electrode_changed(self._on_electrode_changed)

    def _on_electrode_changed(self, data: object) -> None:
        if data is None:
            self._potential_widget.clear_calibration()
        else:
            _name, slope, intercept, unit = data  # type: ignore[misc]
            self._potential_widget.set_calibration(unit, slope, intercept)
        self._potential_widget.refresh()

    def on_close(self) -> None:
        """窗口关闭时清理资源（由顶层调用）。"""
        self._running = False
        self._heartbeat_active = False
        if self._com.is_open:
            self._com.send_reset()
            self._com.shutdown()


__all__ = ["MainWindow", "TitrationState"]
